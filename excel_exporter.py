import os
import pandas as pd
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import config

def export_to_excel(hard_results, soft_results, total_penalty, df_can_bo, df_merged_lich, all_shift_counts, mu):
    """
    Hàm nhận dữ liệu từ main.py và xuất báo cáo ra file Excel.
    """
    print("\n[INFO] Đang tạo báo cáo Excel...")
    wb = openpyxl.Workbook()
    
    # --- Khởi tạo các Style màu sắc ---
    font_title = Font(name='Segoe UI', size=16, bold=True, color='2C3E50')
    font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    fill_pass = PatternFill(start_color='D4EFDF', end_color='D4EFDF', fill_type='solid') # Xanh
    fill_fail = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid') # Đỏ
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # =========================================================================
    # SHEET 1: TỔNG QUAN (SUMMARY)
    # =========================================================================
    ws_sum = wb.active
    ws_sum.title = "Tổng Quan"
    
    ws_sum['A1'] = "BÁO CÁO BENCHMARK LỊCH TRỰC"
    ws_sum['A1'].font = font_title
    
    ws_sum['A3'] = f"TỔNG ĐIỂM PHẠT: {total_penalty:,.2f}"
    ws_sum['A3'].font = Font(size=14, bold=True, color='C0392B')
    
    # Viết bảng Ràng Buộc Cứng
    ws_sum.append([]) # Dòng trống
    ws_sum.append(["Mã Ràng Buộc", "Trạng Thái", "Số Lượng Vi Phạm"])
    for cell in ws_sum[5]:
        cell.font = font_header
        cell.fill = fill_header

    row_idx = 6
    for rb_name, data in hard_results.items():
        ws_sum.cell(row=row_idx, column=1, value=rb_name).border = border_thin
        status_cell = ws_sum.cell(row=row_idx, column=2, value="PASS" if data['pass'] else "FAIL")
        status_cell.fill = fill_pass if data['pass'] else fill_fail
        status_cell.border = border_thin
        
        violation_cell = ws_sum.cell(row=row_idx, column=3, value=data['violations'])
        violation_cell.border = border_thin
        row_idx += 1

    # Viết bảng Ràng Buộc Mềm
    row_idx += 2
    ws_sum.cell(row=row_idx, column=1, value="CHI TIẾT RÀNG BUỘC MỀM").font = Font(bold=True)
    row_idx += 1
    
    ws_sum.cell(row=row_idx, column=1, value="Tên Tiêu Chí").font = font_header
    ws_sum.cell(row=row_idx, column=2, value="Điểm Phạt").font = font_header
    ws_sum.cell(row=row_idx, column=3, value="Mô Tả").font = font_header
    for cell in ws_sum[row_idx]: cell.fill = fill_header
    
    row_idx += 1
    for rb_name, data in soft_results.items():
        ws_sum.cell(row=row_idx, column=1, value=rb_name).border = border_thin
        ws_sum.cell(row=row_idx, column=2, value=data['score']).border = border_thin
        ws_sum.cell(row=row_idx, column=3, value=data['details']).border = border_thin
        row_idx += 1

    # Chỉnh độ rộng cột tự động
    for col in ws_sum.columns:
        col_letter = get_column_letter(col[0].column)
        ws_sum.column_dimensions[col_letter].width = 30

    # =========================================================================
    # LƯU FILE
    # =========================================================================
    # SỬA LẠI KHÚC NÀY:
    os.makedirs("output", exist_ok=True) # Đảm bảo thư mục output tồn tại
    filename = f"Bao_Cao_Benchmark_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    output_path = os.path.join("output", filename)
    
    wb.save(output_path)
    print(f"[SUCCESS] Đã xuất báo cáo thành công ra file: {output_path}")